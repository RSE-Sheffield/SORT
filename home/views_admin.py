"""Custom admin views for superuser portal"""

import csv
import io
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from datetime import timedelta

from home.models import AdminAuditLog, Organisation, Project, User
from survey.models import Survey, SurveyResponse


@staff_member_required
def admin_dashboard(request):
    """Custom admin dashboard with platform statistics"""
    if not request.user.is_superuser:
        raise PermissionDenied("Only superusers can access this dashboard")

    now = timezone.now()
    last_30_days = now - timedelta(days=30)

    stats = {
        "total_users": User.objects.filter(is_superuser=False).count(),
        "total_organisations": Organisation.objects.count(),
        "total_projects": Project.objects.count(),
        "total_surveys": Survey.objects.count(),
        "total_responses": SurveyResponse.objects.count(),
        "consented_surveys": Survey.objects.filter(is_shared=True).count(),
        "consented_responses": SurveyResponse.objects.filter(survey__is_shared=True).count(),
        "new_surveys_30d": Survey.objects.filter(created_at__gte=last_30_days).count(),
        "active_surveys": Survey.objects.filter(is_active=True).count(),
    }

    top_organisations = Organisation.objects.annotate(
        survey_count=Count("projects__survey")
    ).order_by("-survey_count")[:10]

    recent_audit_logs = AdminAuditLog.objects.select_related("performed_by")[:20]

    context = {
        "stats": stats,
        "top_organisations": top_organisations,
        "recent_audit_logs": recent_audit_logs,
    }

    return render(request, "admin/dashboard.html", context)


@staff_member_required
def export_consented_data(request):
    """Export all consented survey data (is_shared=True)"""
    if not request.user.is_superuser:
        raise PermissionDenied()

    export_format = request.GET.get("format", "csv")
    consented_surveys = Survey.objects.filter(is_shared=True).select_related("project__organisation")

    if not consented_surveys.exists():
        messages.warning(request, "No consented surveys available for export")
        return redirect("admin:survey_survey_changelist")

    total_responses = sum(survey.responses_count for survey in consented_surveys)

    if export_format == "csv":
        csv_data = _generate_consented_csv(consented_surveys)

        # Log export
        from home.services import audit_service
        audit_service.log_export(
            user=request.user,
            export_type="CSV",
            survey_count=consented_surveys.count(),
            response_count=total_responses
        )

        response = HttpResponse(csv_data, content_type="text/csv")
        filename = f"sort_consented_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    elif export_format == "excel":
        excel_data = _generate_consented_excel(consented_surveys)

        # Log export
        from home.services import audit_service
        audit_service.log_export(
            user=request.user,
            export_type="Excel",
            survey_count=consented_surveys.count(),
            response_count=total_responses
        )

        response = HttpResponse(
            excel_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"sort_consented_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    messages.error(request, f"Unsupported format: {export_format}")
    return redirect("admin:survey_survey_changelist")


def _generate_consented_csv(surveys):
    """Generate combined CSV using existing Survey.to_csv() pattern"""
    buffer = io.StringIO()

    # Metadata columns + survey fields
    metadata_fields = ["survey_id", "survey_name", "organisation", "project", "response_created_at"]
    all_fields = set()
    for survey in surveys:
        # Skip surveys without valid config
        if survey.survey_config is None:
            continue
        try:
            all_fields.update(survey.fields)
        except (KeyError, TypeError, AttributeError):
            # Survey config is malformed, skip
            continue

    fieldnames = metadata_fields + sorted(all_fields)
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()

    for survey in surveys:
        # Skip surveys without valid config
        if survey.survey_config is None:
            continue

        for response in survey.survey_response.all():
            row_data = {
                "survey_id": survey.pk,
                "survey_name": survey.name,
                "organisation": survey.organisation.name,
                "project": survey.project.name,
                "response_created_at": response.created_at.isoformat(),
            }
            try:
                row_data.update(dict(zip(survey.fields, response.answers_values)))
            except (KeyError, TypeError, AttributeError):
                # Skip malformed responses
                continue
            writer.writerow(row_data)

    return buffer.getvalue()


def _generate_consented_excel(surveys):
    """Generate combined Excel file using existing Survey.to_excel() pattern"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Consented Survey Data"

    # Metadata columns + survey fields
    metadata_fields = ["survey_id", "survey_name", "organisation", "project", "response_created_at"]
    all_fields = set()
    for survey in surveys:
        # Skip surveys without valid config
        if survey.survey_config is None:
            continue
        try:
            all_fields.update(survey.fields)
        except (KeyError, TypeError, AttributeError):
            continue

    headers = metadata_fields + sorted(all_fields)

    # Write headers with bold font
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    row_num = 2
    for survey in surveys:
        # Skip surveys without valid config
        if survey.survey_config is None:
            continue

        for response in survey.survey_response.all():
            row_data = {
                "survey_id": survey.pk,
                "survey_name": survey.name,
                "organisation": survey.organisation.name,
                "project": survey.project.name,
                "response_created_at": response.created_at.isoformat(),
            }
            try:
                row_data.update(dict(zip(survey.fields, response.answers_values)))
            except (KeyError, TypeError, AttributeError):
                continue

            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))

            row_num += 1

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@staff_member_required
def delete_with_reason_view(request, model_name):
    """Intermediate view for deleting objects with audit logging"""
    if not request.user.is_superuser:
        raise PermissionDenied()

    ids_str = request.GET.get("ids", "")
    if not ids_str:
        messages.error(request, "No objects selected")
        return redirect("admin:index")

    ids = [int(id_) for id_ in ids_str.split(",")]

    model_map = {
        "user": User,
        "organisation": Organisation,
        "project": Project,
        "survey": Survey,
    }

    model = model_map.get(model_name)
    if not model:
        messages.error(request, f"Unknown model: {model_name}")
        return redirect("admin:index")

    objects = model.objects.filter(pk__in=ids)

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()

        if not reason:
            messages.error(request, "Reason is required for deletion")
            return render(request, "admin/delete_with_reason.html", {
                "objects": objects,
                "model_name": model_name,
            })

        cascade_info = _calculate_cascade_impact(objects, model)

        # Log each deletion
        from home.services import audit_service
        for obj in objects:
            audit_service.log_deletion(
                user=request.user,
                target=obj,
                reason=reason,
                cascade_info=cascade_info.get(obj.pk, {})
            )

        count = objects.count()
        objects.delete()

        messages.success(request, f"Successfully deleted {count} {model_name}(s)")
        return redirect(f"admin:home_{model_name}_changelist")

    # GET - show confirmation
    cascade_info = _calculate_cascade_impact(objects, model)
    context = {
        "objects": objects,
        "model_name": model_name,
        "cascade_info": cascade_info,
    }

    return render(request, "admin/delete_with_reason.html", context)


def _calculate_cascade_impact(objects, model):
    """Calculate what will be cascade deleted"""
    cascade_map = {}

    if model == Organisation:
        for org in objects:
            cascade_map[org.pk] = {
                "projects": org.projects.count(),
                "surveys": Survey.objects.filter(project__organisation=org).count(),
                "responses": SurveyResponse.objects.filter(survey__project__organisation=org).count(),
            }
    elif model == Project:
        for project in objects:
            cascade_map[project.pk] = {
                "surveys": project.survey.count(),
                "responses": SurveyResponse.objects.filter(survey__project=project).count(),
            }
    elif model == Survey:
        for survey in objects:
            cascade_map[survey.pk] = {
                "responses": survey.survey_response.count(),
            }
    elif model == User:
        for user in objects:
            cascade_map[user.pk] = {
                "organisations": user.organisation_set.count(),
                "projects_created": Project.objects.filter(created_by=user).count(),
            }

    return cascade_map
