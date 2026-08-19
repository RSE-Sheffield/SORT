import io
import openpyxl

from django.test import TestCase

from SORT.test.model_factory import SurveyFactory


class TestSurveyExport(TestCase):
    """
    Test the data export functionality.
    """

    def setUp(self):
        # Set up survey data
        self.survey = SurveyFactory()
        self.survey.initialise()
        self.survey.generate_mock_responses()

    def test_excel_export(self):
        """
        The survey Excel export function should export a workbook containing survey responses.
        """
        # Generate Excel workbook
        buffer = io.BytesIO()
        buffer.write(self.survey.to_excel())
        buffer.seek(0)

        # Read data from Excel file
        workbook = openpyxl.load_workbook(buffer)
        self.assertEqual(len(workbook.sheetnames), 1, "Unexpected sheets found")
        for sheet in workbook.worksheets:
            # Check data shape (rows and columns)
            # Check row count (ignore header row)
            self.assertEqual(sheet.max_row - 1, self.survey.responses_count, "Unexpected row count")
            field_count = len(self.survey.fields)
            for row in sheet.values:
                # The number of cells (columns) in each row should equal the field count
                self.assertEqual(len(row), field_count, "Unexpected column count")

    def test_excel_export_with_long_survey_name(self):
        """
        Excel worksheet names cannot exceed 31 characters. A survey name longer than
        that should be truncated rather than crashing the export.

        Regression test for https://github.com/UniversityOfSheffield/SORT/issues/705
        """
        survey = SurveyFactory(
            name="PILOT - Self-assessment and Organisation Readiness Tool at "
            "MEH FT to build Research Capability and Capacity"
        )
        survey.initialise()
        survey.generate_mock_responses(1)

        buffer = io.BytesIO()
        buffer.write(survey.to_excel())
        buffer.seek(0)

        workbook = openpyxl.load_workbook(buffer)
        self.assertEqual(len(workbook.sheetnames), 1)
        sheet_name = workbook.sheetnames[0]
        self.assertLessEqual(len(sheet_name), 31)
        self.assertEqual(sheet_name, "PILOT - Self-assessment and Org")

    def test_excel_export_with_invalid_sheet_name_characters(self):
        """
        Excel worksheet names cannot contain [ ] : * ? / \\ characters.
        """
        survey = SurveyFactory(name="Q1/Q2 Survey: [Draft]?")
        survey.initialise()
        survey.generate_mock_responses(1)

        buffer = io.BytesIO()
        buffer.write(survey.to_excel())
        buffer.seek(0)

        workbook = openpyxl.load_workbook(buffer)
        sheet_name = workbook.sheetnames[0]
        self.assertEqual(sheet_name, "Q1Q2 Survey Draft")
